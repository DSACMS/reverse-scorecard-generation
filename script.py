import csv
import os
import os.path
import sys
from openpyxl import load_workbook

# CONSTANTS
INPUT_FILES_DIR = "input_files"
IDENTITY_ROWS = (8, 32)  # the rows about Identity start at E8 and end at E32
DEVICES_ROWS = (38, 62)  # etc.
NETWORKS_ROWS = (68, 92)
APPLICATIONS_ROWS = (98, 126)
DATA_ROWS = (132, 160)
CROSS_ROWS = (166, 174)
ALL_ROWS = (
    IDENTITY_ROWS,
    DEVICES_ROWS,
    NETWORKS_ROWS,
    APPLICATIONS_ROWS,
    DATA_ROWS,
    CROSS_ROWS,
)

# check if out.csv exists
if not os.path.isfile("out.csv"):
    # if not, duplicate out_template.csv to out.csv
    with open("out_template.csv", "r") as template_file:
        with open("out.csv", "w") as csv_file:
            csv_file.write(template_file.read())

# append to an existing output CSV file, or create a new one if it doesn't exist
# NOTE: no CSV headers are written
with open("out.csv", "a") as csv_file:
    csv_writer = csv.writer(csv_file, quoting=csv.QUOTE_ALL)

    # loop through all Excel files in the input_files directory
    for file in os.listdir(INPUT_FILES_DIR):
        filename = os.fsdecode(file)

        # skip non-Excel files
        if not filename.endswith(".xlsx"):
            continue

        try:
            # open the Excel file
            wb = load_workbook(
                filename=os.path.join(INPUT_FILES_DIR, filename), read_only=True
            )

            # get QUESTIONNAIRE sheet
            ws = wb["QUESTIONNAIRE"]

            # get HHS system name & ISSO name
            system_name = ws["B3"].value
            if not system_name:
                raise ValueError("System name cell B3 is empty")
            isso_name = ws["B4"].value
            if not isso_name:
                raise ValueError("ISSO name cell B4 is empty")

            # set first few columns
            # Acronym	Component	SubmittedBy	dataCenterEnvironment	UserTypes
            new_csv_row = [system_name, "", isso_name, "", ""]

            # get the data from the sheet and sequentially add it to a list
            for row in ALL_ROWS:
                for i in range(row[0], row[1] + 1, 4):
                    # get current capability score
                    score = ws[f"E{i}"].value

                    # get current capability description
                    # i + score - 1 returns the row number of the capability description
                    new_csv_row.append(ws[f"D{i + score - 1}"].value)

                    # add current capability explanation, and remove newlines
                    new_csv_row.append(
                        ws[f"F{i}"].value.replace("\r", "").replace("\n", " ")
                    )

            # add the data to the CSV output file
            csv_writer.writerow(new_csv_row)

            # close the Excel file
            wb.close()

            # print success message
            print(f"Successfully parsed {filename}")

        except Exception as e:
            print(f"Error when trying to parse {filename}: ", e, file=sys.stderr)
            sys.exit(1)

# print final success message
print("\nAll files have been parsed!")
